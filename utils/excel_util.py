# coding:utf-8
import xlrd
import xlwt
# workbook相关
from openpyxl.workbook import Workbook
# ExcelWriter，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter
# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook


class HandleExcel(object):
    '''Excel相关操作类'''

    def __init__(self):
        self.head_row_labels = ['内容提供商名', '专辑名称', '专辑海报(http)', '看点', '导演', '主演', '搜索关键词', '影片详情', '版权到期日', '地区', '年代',
                                '节目类型', '时长（分钟）', '码流', '视频类型', '专辑类型', '是否3D', '评分', '播放次数', '更新集数', '总集数', '内容名称（剧集）',
                                '内容集数', '期数', '内容看点', '内容类型', '源文件', '码率(bit)', '码率名称', '内容类型2', '源文件2', '码率(bit)2',
                                '码率名称2', '内容类型3', '源文件3', '码率(bit)3', '码率名称3', '内容类型4', '源文件4', '码率(bit)4', '码率名称4',
                                '内容类型5', '源文件5', '码率(bit)5', '码率名称5', '内容类型6', '源文件6', '码率(bit)6', '码率名称6', '内容类型7',
                                '源文件7', '码率(bit)7', '码率名称7', '内容类型8', '源文件8', '码率(bit)8', '码率名称8']

    """
        function：
            读出txt文件中的每一条记录，把它保存在list中
        Param:
            filename:  要读出的文件名
        Return:
            res_list： 返回的记录的list
    """

    def read_from_file(self, filename):
        res_list = []
        file_obj = open(filename, "r", encoding='utf-8')
        for line in file_obj.readlines():
            res_list.append(line)
        file_obj.close()
        return res_list

    """
        function：
            读出*.xlsx中的每一条记录，把它保存在data_dic中返回
        Param:
            excel_name:  要读出的文件名
        Return:
            data_dic： 返回的记录的dict
    """

    def read_excel_with_openpyxl_test(self, excel_name="testexcel2007.xlsx"):
        # 读取excel2007文件
        wb = load_workbook(filename=excel_name)
        # 显示有多少张表
        # print(   "Worksheet range(s):" , wb.defined_names.definedName())
        print("Worksheet name(s):", wb.sheetnames)
        # 取第一张表
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]

        # 显示表名，表行数，表列数
        print("Work Sheet Titile:", ws.title)
        print("Work Sheet Rows:", ws.max_row)
        print("Work Sheet Cols:", ws.max_column)
        # 获取读入的excel表格的有多少行，有多少列
        row_num = ws.max_row
        col_num = ws.max_column
        print("row_num: ", row_num, " col_num: ", col_num)
        # 因为按行，所以返回A1, B1, C1这样的顺序
        for row in ws.rows:
            for cell in row:
                print(cell.value)

        # A1, A2, A3这样的顺序
        for column in ws.columns:
            for cell in column:
                print(cell.value)

        # 建立存储数据的字典
        data_dic = {}
        sign = 1
        # 把数据存到字典中
        for row in ws.rows:
            temp_list = []
            # print( "row",row
            for cell in row:
                print(cell.value,
                      temp_list.append(cell.value))
            print("")
            data_dic[sign] = temp_list
            sign += 1
        print(data_dic)
        return data_dic

    '''
           function：
               保存.xlsx 格式的Excel文件
           Param:
                dataset：要保存的结果数据，list存储
                head_row：excel header信息
                excel_name:  要写入的Excel文件名
           Return:
              无
    '''

    def write_excel_xlsx(self, dataset, head_row, excel_name,header_key=None):
        # 创建文件xlsx
        wb = Workbook()
        wb.save(filename=excel_name)

        wb = load_workbook(filename=excel_name)
        sheet = wb.active

        count = 0
        # 首行标题：
        # 写第一行，标题行 ,row = 1
        if not not head_row:
            for field in range(1, len(head_row) + 1):  # 写入表头
                _ = sheet.cell(row=1, column=field, value=str(head_row[field - 1]))
            # default = self.set_style('Times New Roman', 200, False)  # define style out the loop will work
            count += 1

        # 逐行写入第二行及其以后的那些行
        if not not head_row:
            for line in dataset:
                is_list = line.__class__ == [].__class__
                is_set = line.__class__ == set().__class__
                is_dict = isinstance(line, dict)
                if is_list or is_set:
                    row_list = line
                elif is_dict:
                    row_list = []
                    if not not header_key:
                        for key in header_key:
                            value = ''
                            if key in line:
                                value = line[key]
                            row_list.append(value)
                    else:
                        for k, v in line.items():
                            row_list.append(v)
                else:
                    # .replace(" ", "")  we have space in the dataset
                    row_list = str(line).strip("\n").split("\t")
                count += 1
                for col1 in range(1, len(row_list) + 1):  # 写入数据
                    _ = sheet.cell(row=count, column=col1, value=str(row_list[col1 - 1]))

        wb.save(filename=excel_name)
        return

    '''
        function：
            读出.xlsx 格式的Excel文件
        Param:
            excel_name:  要读出的Excel文件名
        Return:
           无
    '''

    def read_excel_xlsx(self, excel_name):
        # 读取excel2007文件
        wb = load_workbook(filename=excel_name)
        # 显示有多少张表
        # print(   "Worksheet range(s):" , wb.defined_names.definedName())
        print("Worksheet name(s):", wb.sheetnames)
        # 取第一张表
        sheetnames = wb.sheetnames
        ws = wb[sheetnames[0]]

        # 显示表名，表行数，表列数
        '''
        print(   "Work Sheet Titile:" ,ws.title)
        print(   "Work Sheet Rows:" ,ws.max_row)
        print(   "Work Sheet Cols:" ,ws.max_column)
        # 获取读入的excel表格的有多少行，有多少列
        row_num=ws.max_row
        col_num=ws.max_column
        print( "row_num: ",row_num," col_num: ",col_num)
        # 因为按行，所以返回A1, B1, C1这样的顺序
        for row in ws.rows:
            for cell in row:
                print(cell.value)
        # A1, A2, A3这样的顺序
        for column in ws.columns:
            for cell in column:
                print(cell.value)
        '''
        # ws.row_values()
        return ws

    """
            function：
                读出Excel文件
            Param:
                excel_name:  要读出的Excel文件名
            Return:
               无
        """

    def read_excel(self, excel_name):
        workbook = xlrd.open_workbook(excel_name)
        # 获取所有sheet
        # print( workbook.sheet_names() )# [u'sheet1', u'sheet2']
        sheet1 = workbook.sheet_by_index(0)  # sheet索引从0开始
        # sheet的名称，行数，列数
        # print( sheet1.name,"rows =",sheet1.nrows,"clos=",sheet1.ncols)
        return sheet1

    """
        function：
            测试输出Excel内容
            读出Excel文件
        Param:
            excel_name:  要读出的Excel文件名
        Return:
           无
    """

    def read_excel_test(self, excel_name):
        workbook = xlrd.open_workbook(excel_name)
        # 获取所有sheet
        print(workbook.sheet_names())  # [u'sheet1', u'sheet2']
        # sheet2_name = workbook.sheet_names()[1]
        # 根据sheet索引或者名称获取sheet内容
        sheet1 = workbook.sheet_by_index(0)  # sheet索引从0开始
        # sheet2 = workbook.sheet_by_name('Sheet1')
        # sheet的名称，行数，列数
        print(sheet1.name, "rows =", sheet1.nrows, "clos=", sheet1.ncols)
        # 获取整行和整列的值（数组）
        for row_num in range(0, sheet1.nrows):
            rows = sheet1.row_values(row_num)
            # print("row",row_num,"[0]",rows[0])
            print("row", row_num, "[0]", "http://f666666.xyz/video/banyungong" + rows[0].split('/')[-1])

        rows = sheet1.row_values(3)  # 获取第四行内容
        cols = sheet1.col_values(2)  # 获取第三列内容
        print(rows)
        print(cols)
        # 获取单元格内容
        print(sheet1.cell(1, 0).value)
        print(sheet1.cell_value(1, 0))
        print(sheet1.row(1)[0].value)
        # 获取单元格内容的数据类型
        print(sheet1.cell(1, 0).ctype)
        # 通过名称获取
        return workbook.sheet_by_index(0)

    """
        function：
            设置单元格样式
        Param:
            name:  字体名字
            height:  字体高度
            bold:  是否大写
        Return:
            style: 返回设置好的格式对象
    """

    def set_style(self, name, height, bold=False):
        style = xlwt.XFStyle()  # 初始化样式
        font = xlwt.Font()  # 为样式创建字体
        font.name = name  # 'Times New Roman'
        font.bold = bold
        font.color_index = 4
        font.height = height
        borders = xlwt.Borders()
        borders.left = 6
        borders.right = 6
        borders.top = 6
        borders.bottom = 6
        style.font = font
        style.borders = borders
        return style

    """
        function：
            按照 设置单元格样式  把计算结果由txt转变为Excel存储
        Param:
            dataset：要保存的结果数据，list存储
        Return:
            将结果保存为 excel对象中
    """

    def write_to_excel(self, dataset, head_row, save_excel_name, header_key=None):
        f = xlwt.Workbook()  # 创建工作簿
        # 创建第一个sheet:
        # sheet1
        count = 1
        sheet1 = f.add_sheet(u'Sheet1', cell_overwrite_ok=True)  # 创建sheet
        # 首行标题：
        for p in range(len(head_row)):
            sheet1.write(0, p, head_row[p], self.set_style('Times New Roman', 250, True))
        default = self.set_style('Times New Roman', 200, False)  # define style out the loop will work
        for line in dataset:
            is_list = line.__class__ == [].__class__
            is_set = line.__class__ == set().__class__
            is_dict = isinstance(line, dict)
            row_list = None
            if is_list or is_set:
                row_list = line
            elif is_dict:
                row_list = []
                if not not header_key:
                    for key in header_key:
                        value = ''
                        if key in line:
                            value = line[key]
                        row_list.append(value)
                else:
                    for k, v in line.items():
                        row_list.append(v)
            else:
                # .replace(" ", "")  we have space in the dataset
                row_list = str(line).strip("\n").split("\t")
            for pp in range(len(row_list)):
                sheet1.write(count, pp, row_list[pp], default)
            count += 1

        f.save(save_excel_name)  # 保存文件

    # 暂时无法使用
    def run_main_save_to_excel_with_openpyxl(self):
        print("测试读写2007及以后的excel文件xlsx，以方便写入文件更多数据")
        print("1. 把txt文件读入到内存中，以list对象存储")
        dataset_list = self.read_from_file("../test/data/test_excel.txt")
        '''test use openpyxl to handle EXCEL 2007'''
        print("2. 把文件写入到Excel表格中")
        head_row_label = self.head_row_labels
        save_name = "test_openpyxl.xlsx"
        self.write_to_excel_with_openpyxl(dataset_list, head_row_label, save_name)
        print("3.  执行完毕，由txt格式文件保存为Excel文件的任务")

    def run_main_save_to_excel_with_xlwt(self):
        print(" 4. 把txt文件读入到内存中，以list对象存储")
        dataset_list = self.read_from_file("../test/data/test_excel.txt")
        '''test use xlwt to handle EXCEL 97-2003'''
        print(" 5. 把文件写入到Excel表格中")
        head_row_label = self.head_row_labels
        save_name = "test_xlwt.xls"
        self.write_to_excel(dataset_list, head_row_label, save_name)
        print("6. 执行完毕，由txt格式文件保存为Excel文件的任务")


def util_excel_test():
    print("create handle Excel Object")
    obj_handle_excel = HandleExcel()
    # 分别使用openpyxl和xlwt将数据写入文件
    obj_handle_excel.run_main_save_to_excel_with_openpyxl()
    obj_handle_excel.run_main_save_to_excel_with_xlwt()
    '''测试读出文件，注意openpyxl不可以读取xls的文件,xlrd不可以读取xlsx格式的文件'''
    # obj_handle_excel.read_excel_with_openpyxl("testexcel2003.xls")  # 错误写法
    # obj_handle_excel.read_excel("test_xlwt.xls")
    # 暂时无法使用
    obj_handle_excel.read_excel_with_openpyxl_test("../test/data/test.xlsx")


if __name__ == '__main__':
    # util_excel_test()
    sheet = HandleExcel().read_excel_xlsx('ISO 639 lang.xlsx')
    output = 'ISO_639_lang.json'
    try:
        ISO_639_lang_map = {}
        row_num = 0
        for row in sheet.rows:
            row_num += 1
            if row_num == 1:
                continue

            row_value = row
            key = row_value[1].value
            ISO_639_lang_map[key] = {}
            ISO_639_lang_map[key]['iso'] = row_value[0].value
            ISO_639_lang_map[key]['iso639_1'] = row_value[1].value
            ISO_639_lang_map[key]['iso639_2'] = row_value[2].value
            print("row {} iso ={} ,iso639_1= {},iso639_2={}".format(row_num, row_value[0].value, row_value[1].value,
                                                                    row_value[2].value))
        # load org data from db file
        import json

        with open(output, 'w', encoding='utf-8') as f:
            json.dump(ISO_639_lang_map, f)
            f.close()

    except BaseException as b:
        print(b)
